"""
NexERP Multi-Tab Excel & CSV Export Service.
Generates structured financial and operational workbooks using openpyxl and standard CSV.
"""

import io
import csv
from typing import Any, Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExportService:
    """
    Spreadsheet and CSV generator.
    """

    @classmethod
    def generate_csv(cls, headers: List[str], rows: List[List[Any]]) -> str:
        """Generate formatted CSV string."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        return output.getvalue()

    @classmethod
    def generate_financial_excel_workbook(
        cls,
        sheet_title: str,
        headers: List[str],
        data_rows: List[List[Any]],
        company_name: str = "Apex Dynamics Enterprise"
    ) -> bytes:
        """
        Generate enterprise-styled Excel workbook (.xlsx).
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title

        # Header styling
        title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # Title Rows
        ws.cell(row=1, column=1, value=company_name).font = title_font
        ws.cell(row=2, column=1, value=f"{sheet_title} Report").font = Font(name="Calibri", size=12, italic=True)

        # Table Column Headers
        start_row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data Rows
        for r_idx, row_values in enumerate(data_rows, start=start_row + 1):
            for c_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
